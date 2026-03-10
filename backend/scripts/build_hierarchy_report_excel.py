# -*- coding: utf-8 -*-
"""
Пример отчётности по иерархии и всем отделам для Медхаус / Свисс Энерджи.
Иерархия: Руководство → Компании → Регионы → Подразделения → Детализация по отделам.
Запуск: python -m scripts.build_hierarchy_report_excel
Результат: PROJECT_ROOT / Отчетность_по_иерархии_пример.xlsx
"""
import sys
from pathlib import Path

BACKEND_DIR = Path(__file__).resolve().parent.parent
PROJECT_ROOT = BACKEND_DIR.parent
sys.path.insert(0, str(BACKEND_DIR))

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

OUTPUT_PATH = PROJECT_ROOT / "Отчетность_по_иерархии_пример.xlsx"

HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _write_header(ws, row, headers):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = HEADER_FONT_WHITE
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = THIN_BORDER


def _write_data(ws, start_row, rows, border=True):
    for i, row_data in enumerate(rows):
        r = start_row + i
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=r, column=col, value=val)
            if border:
                c.border = THIN_BORDER


def sheet_oglavlenie(wb):
    """Оглавление: иерархия отчётности и список листов."""
    ws = wb.create_sheet("Оглавление и иерархия", 0)
    ws.merge_cells("A1:D1")
    ws["A1"] = "Отчётность по иерархии — Медхаус / Свисс Энерджи (пример)"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Уровень 1 — Руководство: сводка для руководства (KPI по компаниям)."
    ws["A3"] = "Уровень 2 — По компаниям: Медхаус и Свисс Энерджи отдельно."
    ws["A4"] = "Уровень 3 — По регионам: филиалы и города."
    ws["A5"] = "Уровень 4 — По подразделениям: MedPack, HomeTest, Dr.Frei, ТП, Логистика, Финансы."
    ws["A6"] = "Уровень 5 — Детализация: продажи, план-факт, товары, логистика, касса/банк, мотивация."
    for r in range(2, 7):
        ws.cell(row=r, column=1).font = Font(size=11)
    row = 9
    ws.cell(row=row, column=1, value="Листы в этой книге:").font = HEADER_FONT
    row += 1
    sheets_desc = [
        ("Уровень 1 — Руководство", "Сводные KPI для руководства"),
        ("Уровень 2 — По компаниям", "Выручка и показатели по Медхаус / Свисс Энерджи"),
        ("Уровень 3 — По регионам", "Показатели по городам и филиалам"),
        ("Уровень 4 — По подразделениям", "MedPack, HomeTest, Dr.Frei, ТП, Логистика, Финансы"),
        ("Уровень 5 — Продажи", "Детализация продаж по периодам и филиалам"),
        ("Уровень 5 — План-факт", "План-факт по отделам/регионам"),
        ("Уровень 5 — Товары", "Номенклатура по подразделениям"),
        ("Уровень 5 — Логистика", "Остатки по складам"),
        ("Уровень 5 — Касса и банк", "Движение денег по городам"),
        ("Мотивация ТП", "План/факт отгрузок и премии по ТП"),
        ("Дашборд сводный", "Графики и ключевые показатели"),
    ]
    for name, desc in sheets_desc:
        ws.cell(row=row, column=1, value=name).font = Font(bold=True)
        ws.cell(row=row, column=2, value=desc)
        row += 1
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 50


def sheet_uroven1_rukovodstvo(wb):
    """Уровень 1 — для руководства: сводные KPI по компаниям."""
    ws = wb.create_sheet("Уровень 1 — Руководство", 1)
    ws["A1"] = "Сводка для руководства"
    ws["A1"].font = Font(bold=True, size=12)
    _write_header(ws, 3, ["Компания", "Выручка (тыс. тг)", "Выполнение плана %", "Городов", "Период"])
    _write_data(ws, 4, [
        ["Медхаус", 26800, 98.5, 15, "2025"],
        ["Свисс Энерджи", 4200, 102.0, 3, "2025"],
        ["Итого", 31000, 99.0, 18, "2025"],
    ])
    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 22


def sheet_uroven2_kompanii(wb):
    """Уровень 2 — по компаниям."""
    ws = wb.create_sheet("Уровень 2 — По компаниям", 2)
    ws["A1"] = "Отчётность по компаниям"
    ws["A1"].font = Font(bold=True, size=12)
    _write_header(ws, 3, ["Компания", "Год", "Выручка (тыс. тг)", "Прибыль (тыс. тг)", "Записей"])
    _write_data(ws, 4, [
        ["Медхаус", 2023, 2150, 420, 520],
        ["Медхаус", 2024, 2420, 480, 580],
        ["Медхаус", 2025, 2680, 530, 610],
        ["Свисс Энерджи", 2023, 380, 75, 120],
        ["Свисс Энерджи", 2024, 420, 82, 135],
        ["Свисс Энерджи", 2025, 450, 90, 145],
    ])
    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 20


def sheet_uroven3_regiony(wb):
    """Уровень 3 — по регионам/филиалам."""
    ws = wb.create_sheet("Уровень 3 — По регионам", 3)
    ws["A1"] = "Отчётность по регионам и филиалам"
    ws["A1"].font = Font(bold=True, size=12)
    _write_header(ws, 3, ["Регион / Город", "Компания", "Выручка (тыс. тг)", "Доля %", "Филиалов"])
    _write_data(ws, 4, [
        ["Алматы", "Медхаус", 18500, 69.0, 1],
        ["Астана", "Медхаус", 15100, 56.3, 1],
        ["Шымкент", "Медхаус", 10500, 39.2, 1],
        ["Актобе", "Медхаус", 4200, 15.7, 1],
        ["Караганда", "Медхаус", 3800, 14.2, 1],
        ["Атырау", "Медхаус", 2100, 7.8, 1],
        ["Уральск", "Медхаус", 1800, 6.7, 1],
        ["Прочие", "Медхаус", 2800, 10.4, 7],
        ["Итого Медхаус", "", 26800, 100, 15],
        ["Свисс Энерджи (сводно)", "Свисс Энерджи", 4500, 100, 3],
    ])
    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 22


def sheet_uroven4_podrazdeleniya(wb):
    """Уровень 4 — по подразделениям (отделам)."""
    ws = wb.create_sheet("Уровень 4 — По подразделениям", 4)
    ws["A1"] = "Отчётность по подразделениям (отделам)"
    ws["A1"].font = Font(bold=True, size=12)
    _write_header(ws, 3, ["Подразделение / Отдел", "Описание", "Выручка (тыс. тг)", "Доля %", "Регионов"])
    _write_data(ws, 4, [
        ["MedPack", "Презервативы, упаковка", 18500, 69.0, 15],
        ["HomeTest", "Экспресс-тесты", 6200, 23.1, 12],
        ["Dr.Frei", "Медтехника (тонометры, ингаляторы)", 2100, 7.8, 8],
        ["Торговые представители (ТП)", "Продажи в регионах", 26800, 100, 15],
        ["Логистика", "Склады, остатки", "—", "—", 8],
        ["Финансы (Касса/Банк)", "Поступления по городам", "—", "—", 15],
    ])
    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 28


def sheet_uroven5_prodazhi(wb):
    """Уровень 5 — продажи (детализация)."""
    ws = wb.create_sheet("Уровень 5 — Продажи", 5)
    _write_header(ws, 1, ["Период", "Регион", "Подразделение", "Сумма (тг)", "Кол-во"])
    _write_data(ws, 2, [
        ["2025-01", "Алматы", "MedPack", 18500000, 1200],
        ["2025-01", "Астана", "MedPack", 14200000, 980],
        ["2025-01", "Шымкент", "HomeTest", 9800000, 650],
        ["2025-02", "Алматы", "MedPack", 19200000, 1250],
        ["2025-02", "Астана", "Dr.Frei", 15100000, 1020],
        ["2025-03", "Шымкент", "MedPack", 10500000, 710],
    ])
    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 20


def sheet_uroven5_plan_fact(wb):
    """Уровень 5 — план-факт по отделам."""
    ws = wb.create_sheet("Уровень 5 — План-факт", 6)
    _write_header(ws, 1, ["Отдел / Регион", "Период", "План (тыс. тг)", "Факт (тыс. тг)", "Выполнение %"])
    _write_data(ws, 2, [
        ["MedPack", "2025-01", 19000, 18500, 97.4],
        ["MedPack", "2025-02", 19500, 19200, 98.5],
        ["HomeTest", "2025-01", 6500, 6200, 95.4],
        ["HomeTest", "2025-02", 6600, 6500, 98.5],
        ["ТП Алматы", "2025-01", 20000, 19800, 99.0],
        ["ТП Астана", "2025-01", 15000, 15100, 100.7],
    ])
    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 22


def sheet_uroven5_tovary(wb):
    """Уровень 5 — товары по подразделениям."""
    ws = wb.create_sheet("Уровень 5 — Товары", 7)
    _write_header(ws, 1, ["Подразделение", "Наименование", "Кол-во", "Сумма (тг)", "Период"])
    _write_data(ws, 2, [
        ["MedPack", "DOLPHI 3 в 1 №12", 12500, 18500000, "2025"],
        ["MedPack", "DOLPHI 3 в 1 №3", 9800, 9200000, "2025"],
        ["HomeTest", "Экспресс-тест беременность №2", 8200, 12300000, "2025"],
        ["HomeTest", "Экспресс-тест овуляция №5", 5400, 7600000, "2025"],
        ["Dr.Frei", "Gamma Control тонометр", 2100, 9800000, "2025"],
        ["Dr.Frei", "Ингалятор Gamma NEMO", 850, 4200000, "2025"],
    ])
    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 32


def sheet_uroven5_logistika(wb):
    """Уровень 5 — логистика по складам."""
    ws = wb.create_sheet("Уровень 5 — Логистика", 8)
    _write_header(ws, 1, ["Склад", "Товар", "Приход", "Расход", "Остаток", "Ед."])
    _write_data(ws, 2, [
        ["Алматы", "DOLPHI 3 в 1 №12", 5000, 4200, 800, "шт"],
        ["Астана", "DOLPHI 3 в 1 №12", 3200, 3000, 200, "шт"],
        ["Шымкент", "Gamma Control", 150, 140, 10, "шт"],
        ["Актобе", "DOLPHI 3 в 1 №3", 2100, 1950, 150, "шт"],
    ])
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 24


def sheet_uroven5_kassa_bank(wb):
    """Уровень 5 — касса и банк по городам."""
    ws = wb.create_sheet("Уровень 5 — Касса и банк", 9)
    _write_header(ws, 1, ["Дата", "Город", "Касса (тг)", "Банк (тг)", "Тип оплаты"])
    _write_data(ws, 2, [
        ["2026-02-01", "Шымкент", 0, 16805, "Без нал"],
        ["2026-02-02", "Тараз", 50000, 0, "Нал"],
        ["2026-02-02", "Шымкент", 170010, 0, "Нал"],
        ["2026-02-28", "Уральск", 0, 283906, "Без нал"],
        ["2026-02-28", "Шымкент", 0, 748620, "Без нал"],
    ])
    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 18


def sheet_motivaciya(wb):
    """Мотивация ТП по отделам/городам."""
    ws = wb.create_sheet("Мотивация ТП", 10)
    _write_header(ws, 1, ["Город", "ФИО", "Оклад", "План отгрузки", "Факт отгрузки", "Коэф.", "Премия", "Итого"])
    _write_data(ws, 2, [
        ["Актау", "Иванов И.И.", 40000, 10179841, 8354913, 0.82, 121376, 121376],
        ["Актобе", "Петров П.П.", 40000, 8583795, 6928007, 0.81, 92871, 92871],
        ["Алматы", "Сидорова С.С.", 40000, 12500000, 11800000, 0.94, 83000, 83000],
        ["Шымкент", "Козлов К.К.", 40000, 9800000, 9250000, 0.94, 77000, 77000],
    ])
    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 16


def sheet_dashboard(wb):
    """Дашборд сводный — ключевые цифры."""
    ws = wb.create_sheet("Дашборд сводный", 11)
    ws["A1"] = "Дашборд — ключевые показатели"
    ws["A1"].font = Font(bold=True, size=14)
    _write_data(ws, 3, [
        ["Показатель", "Значение", "Период"],
        ["Выручка всего (тыс. тг)", 31000, "2025"],
        ["Выполнение плана (%)", 99.0, "2025"],
        ["Активных городов", 15, "на дату"],
        ["Подразделений в отчётности", 6, "—"],
    ], border=True)
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 18


def main():
    wb = Workbook()
    wb.remove(wb.active)
    sheet_oglavlenie(wb)
    sheet_uroven1_rukovodstvo(wb)
    sheet_uroven2_kompanii(wb)
    sheet_uroven3_regiony(wb)
    sheet_uroven4_podrazdeleniya(wb)
    sheet_uroven5_prodazhi(wb)
    sheet_uroven5_plan_fact(wb)
    sheet_uroven5_tovary(wb)
    sheet_uroven5_logistika(wb)
    sheet_uroven5_kassa_bank(wb)
    sheet_motivaciya(wb)
    sheet_dashboard(wb)
    wb.save(OUTPUT_PATH)
    print("Created:", str(OUTPUT_PATH))
    print("Sheets: Oglavlenie, Uroven 1-5, Motivaciya TP, Dashboard")


if __name__ == "__main__":
    main()
