# -*- coding: utf-8 -*-
"""
Создаёт один Excel-файл со всеми нужными таблицами и примерным дашбордом.
Запуск: python -m scripts.build_dashboard_excel
Результат: PROJECT_ROOT / Дашборд_управленческий_сводный.xlsx
"""
import sys
from pathlib import Path

# корень backend
BACKEND_DIR = Path(__file__).resolve().parent.parent
PROJECT_ROOT = BACKEND_DIR.parent
sys.path.insert(0, str(BACKEND_DIR))

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# выходной файл
OUTPUT_PATH = PROJECT_ROOT / "Дашборд_управленческий_сводный.xlsx"

# стили
HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _write_header(ws, row, headers):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = HEADER_FONT_WHITE
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = THIN_BORDER
    return row


def _write_data_rows(ws, start_row, rows, border=True):
    for i, row_data in enumerate(rows):
        r = start_row + i
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=r, column=col, value=val)
            if border:
                c.border = THIN_BORDER
    return start_row + len(rows)


def build_sheet_summary(wb):
    """Сводка по годам (выручка, кол-во записей)."""
    ws = wb.create_sheet("Сводка по годам", 0)
    headers = ["Год", "Выручка (тыс. тг)", "Кол-во записей", "Примечание"]
    _write_header(ws, 1, headers)
    data = [
        [2020, 1245, 320, "Анализ 2"],
        [2021, 1582, 410, "Анализ 2"],
        [2022, 1890, 480, "Анализ 2"],
        [2023, 2150, 520, "Анализ 2"],
        [2024, 2420, 580, "Анализ 2"],
        [2025, 2680, 610, "Анализ 2"],
    ]
    _write_data_rows(ws, 2, data)
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 14
    return "Сводка по годам"


def build_sheet_sales(wb):
    """Продажи по месяцам/филиалам (пример)."""
    ws = wb.create_sheet("Продажи по месяцам", 1)
    headers = ["Период", "Филиал / Город", "Сумма (тг)", "Количество", "Валюта"]
    _write_header(ws, 1, headers)
    data = [
        ["2025-01", "Алматы", 18500000, 1200, "KZT"],
        ["2025-01", "Астана", 14200000, 980, "KZT"],
        ["2025-01", "Шымкент", 9800000, 650, "KZT"],
        ["2025-02", "Алматы", 19200000, 1250, "KZT"],
        ["2025-02", "Астана", 15100000, 1020, "KZT"],
        ["2025-02", "Шымкент", 10100000, 680, "KZT"],
        ["2025-03", "Алматы", 19800000, 1280, "KZT"],
        ["2025-03", "Астана", 15800000, 1050, "KZT"],
        ["2025-03", "Шымкент", 10500000, 710, "KZT"],
    ]
    _write_data_rows(ws, 2, data)
    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 18
    return "Продажи по месяцам"


def build_sheet_products(wb):
    """Топ товары по сумме и количеству."""
    ws = wb.create_sheet("Топ товары", 2)
    headers = ["Артикул", "Наименование", "Группа", "Количество", "Сумма (тг)", "Период"]
    _write_header(ws, 1, headers)
    data = [
        ["DOLPHI-12", "DOLPHI 3 в 1 №12 презервативы", "MedPack", 12500, 18500000, "2025"],
        ["HT-02", "Экспресс-тест на беременность №2", "HomeTest", 8200, 12300000, "2025"],
        ["GF-01", "Gamma Control тонометр", "Dr.Frei", 2100, 9800000, "2025"],
        ["DOLPHI-03", "DOLPHI 3 в 1 №3 презервативы", "MedPack", 9800, 9200000, "2025"],
        ["HT-05", "Экспресс-тест на овуляцию №5", "HomeTest", 5400, 7600000, "2025"],
        ["GF-NEMO", "Ингалятор Gamma NEMO", "Dr.Frei", 850, 4200000, "2025"],
    ]
    _write_data_rows(ws, 2, data)
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 22
    return "Топ товары"


def build_sheet_plan_fact(wb):
    """План-факт по месяцам года."""
    ws = wb.create_sheet("План-факт", 3)
    headers = ["Период", "План (тыс. тг)", "Факт (тыс. тг)", "Выполнение %", "Отклонение", "Комментарий"]
    _write_header(ws, 1, headers)
    data = [
        ["2025-01", 45000, 42500, 94.4, -2500, ""],
        ["2025-02", 46000, 44300, 96.3, -1700, ""],
        ["2025-03", 47000, 46100, 98.1, -900, ""],
        ["2025-04", 48000, 47200, 98.3, -800, ""],
        ["2025-05", 49000, 48500, 99.0, -500, ""],
        ["2025-06", 50000, 49800, 99.6, -200, ""],
        ["2025-07", 51000, 50200, 98.4, -800, ""],
        ["2025-08", 52000, 51800, 99.6, -200, ""],
        ["2025-09", 53000, 52500, 99.1, -500, ""],
        ["2025-10", 54000, 53800, 99.6, -200, ""],
        ["2025-11", 55000, 54500, 99.1, -500, ""],
        ["2025-12", 56000, 55200, 98.6, -800, ""],
    ]
    _write_data_rows(ws, 2, data)
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 16
    return "План-факт"


def build_sheet_regions(wb):
    """Регионы (пример — города присутствия)."""
    ws = wb.create_sheet("Регионы", 4)
    headers = ["Код региона (KATO)", "Наименование", "Родитель (id)", "Показатель", "Значение"]
    _write_header(ws, 1, headers)
    data = [
        [1, "Алматы", 0, "Выручка тыс.", 18500],
        [2, "Астана", 0, "Выручка тыс.", 15100],
        [3, "Шымкент", 0, "Выручка тыс.", 10500],
        [4, "Актобе", 0, "Выручка тыс.", 4200],
        [5, "Караганда", 0, "Выручка тыс.", 3800],
        [6, "Атырау", 0, "Выручка тыс.", 2100],
        [7, "Уральск", 0, "Выручка тыс.", 1800],
    ]
    _write_data_rows(ws, 2, data)
    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 20
    return "Регионы"


def build_sheet_motivation(wb):
    """Мотивация ТП (план/факт отгрузок, премия) — пример."""
    ws = wb.create_sheet("Мотивация ТП", 5)
    headers = [
        "№", "Организация", "Город", "ФИО", "Оклад (тг)", "План отгрузки", "Факт отгрузки",
        "Коэф.", "Премия отгрузка", "Премия возврат", "Итого к получению",
    ]
    _write_header(ws, 1, headers)
    data = [
        [1, "MedHouse", "Актау", "Иванов И.И.", 40000, 10179841, 8354913, 0.82, 45116, 76260, 121376],
        [2, "MedHouse", "Актобе", "Петров П.П.", 40000, 8583795, 6928007, 0.81, 37411, 55460, 92871],
        [3, "MedHouse", "Алматы", "Сидорова С.С.", 40000, 12500000, 11800000, 0.94, 52000, 31000, 83000],
        [4, "MedHouse", "Шымкент", "Козлов К.К.", 40000, 9800000, 9250000, 0.94, 48000, 29000, 77000],
    ]
    _write_data_rows(ws, 2, data)
    for col in range(1, 12):
        ws.column_dimensions[get_column_letter(col)].width = 14
    return "Мотивация ТП"


def build_sheet_cash_bank(wb):
    """Касса и банк — свод по городам (пример)."""
    ws = wb.create_sheet("Касса и банк", 6)
    headers = ["Дата", "Город", "Касса (тг)", "Банк (тг)", "Тип оплаты", "Контрагент"]
    _write_header(ws, 1, headers)
    data = [
        ["2026-02-01", "Шымкент", 0, 16805, "Без нал", "Al-KausarPharm"],
        ["2026-02-01", "Уральск", 0, 27900, "Без нал", "Talapovna ИП"],
        ["2026-02-02", "Тараз", 50000, 0, "Нал", "Юшкина ИП"],
        ["2026-02-02", "Шымкент", 40040, 0, "Нал", "СУХРАБЖАН ИП"],
        ["2026-02-02", "Шымкент", 130970, 0, "Нал", "СУХРАБЖАН ИП"],
        ["2026-02-27", "Караганда", 30500, 0, "Нал", "Ульянова ИП"],
        ["2026-02-27", "Петропавловск", 50185, 0, "Нал", "В и В ТОО"],
        ["2026-02-28", "Уральск", 0, 283906, "Без нал", ""],
        ["2026-02-28", "Шымкент", 0, 748620, "Без нал", ""],
    ]
    _write_data_rows(ws, 2, data)
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 16
    return "Касса и банк"


def build_sheet_logistics(wb):
    """Логистика — остатки по складам (пример)."""
    ws = wb.create_sheet("Логистика", 7)
    headers = ["Дата", "Склад", "Товар / Номенклатура", "Приход", "Расход", "Остаток", "Ед. изм."]
    _write_header(ws, 1, headers)
    data = [
        ["2026-02", "Алматы", "DOLPHI 3 в 1 №12", 5000, 4200, 800, "шт"],
        ["2026-02", "Алматы", "Экспресс-тест беременность", 3000, 2800, 200, "шт"],
        ["2026-02", "Астана", "DOLPHI 3 в 1 №12", 3200, 3000, 200, "шт"],
        ["2026-02", "Шымкент", "Gamma Control", 150, 140, 10, "шт"],
        ["2026-02", "Актобе", "DOLPHI 3 в 1 №3", 2100, 1950, 150, "шт"],
    ]
    _write_data_rows(ws, 2, data)
    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 22
    return "Логистика"


def build_sheet_dashboard(wb):
    """Лист «Дашборд» — сводные цифры и графики."""
    ws = wb.create_sheet("Дашборд", 8)
    # Заголовок
    ws.merge_cells("A1:F1")
    ws["A1"] = "Управленческий дашборд — Медхаус / Свисс Энерджи"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 24

    # Ключевые показатели (примеры — можно заменить на формулы)
    ws["A3"] = "Ключевые показатели"
    ws["A3"].font = HEADER_FONT
    kpi = [
        ["Показатель", "Значение", "Период"],
        ["Выручка за год (тыс. тг)", 26800, "2025"],
        ["Выполнение плана (%)", 98.5, "2025"],
        ["Кол-во активных городов", 15, "на дату"],
        ["Кол-во записей в базе", 610, "2025"],
    ]
    row = 4
    for r in kpi:
        for c, v in enumerate(r, 1):
            ws.cell(row=row, column=c, value=v).border = THIN_BORDER
        row += 1

    # Таблица для графика «Динамика по годам» (дублируем данные с листа Сводка)
    row += 2
    ws.cell(row=row, column=1, value="Динамика выручки по годам (тыс. тг)")
    ws.cell(row=row, column=1).font = HEADER_FONT
    row += 1
    _write_header(ws, row, ["Год", "Выручка (тыс. тг)"])
    row += 1
    chart_data_start = row
    for i, (y, v) in enumerate([(2020, 1245), (2021, 1582), (2022, 1890), (2023, 2150), (2024, 2420), (2025, 2680)]):
        ws.cell(row=row + i, column=1, value=y)
        ws.cell(row=row + i, column=2, value=v)
    chart_data_end = row + 5
    row = chart_data_end + 2

    # График: столбчатая диаграмма по годам
    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = "Выручка по годам"
    chart1.y_axis.title = "тыс. тг"
    chart1.x_axis.title = "Год"
    data = Reference(ws, min_col=2, min_row=chart_data_start, max_row=chart_data_end)
    cats = Reference(ws, min_col=1, min_row=chart_data_start, max_row=chart_data_end)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.width = 14
    chart1.height = 8
    ws.add_chart(chart1, f"A{row}")

    # План-факт по месяцам (таблица для второго графика)
    row += 18
    ws.cell(row=row, column=1, value="План-факт по месяцам 2025 (тыс. тг)")
    ws.cell(row=row, column=1).font = HEADER_FONT
    row += 1
    _write_header(ws, row, ["Месяц", "План", "Факт"])
    row += 1
    pf_start = row
    months = ["янв", "фев", "мар", "апр", "май", "июн", "июл", "авг", "сен", "окт", "ноя", "дек"]
    plan_vals = [45000, 46000, 47000, 48000, 49000, 50000, 51000, 52000, 53000, 54000, 55000, 56000]
    fact_vals = [42500, 44300, 46100, 47200, 48500, 49800, 50200, 51800, 52500, 53800, 54500, 55200]
    for i in range(12):
        ws.cell(row=row + i, column=1, value=months[i])
        ws.cell(row=row + i, column=2, value=plan_vals[i])
        ws.cell(row=row + i, column=3, value=fact_vals[i])
    pf_end = row + 11
    row = pf_end + 2

    # График: линейный План vs Факт
    chart2 = LineChart()
    chart2.style = 12
    chart2.title = "План и факт по месяцам"
    chart2.y_axis.title = "тыс. тг"
    chart2.x_axis.title = "Месяц"
    data_plan = Reference(ws, min_col=2, min_row=pf_start - 1, max_row=pf_end)
    data_fact = Reference(ws, min_col=3, min_row=pf_start - 1, max_row=pf_end)
    cats_pf = Reference(ws, min_col=1, min_row=pf_start, max_row=pf_end)
    chart2.add_data(data_plan, titles_from_data=True)
    chart2.add_data(data_fact, titles_from_data=True)
    chart2.set_categories(cats_pf)
    chart2.width = 14
    chart2.height = 8
    # второй график справа от первого
    ws.add_chart(chart2, f"J12")

    # Ширина колонок
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14
    return "Дашборд"


def main():
    wb = Workbook()
    # удаляем дефолтный лист после добавления первого своего
    wb.remove(wb.active)

    build_sheet_summary(wb)
    build_sheet_sales(wb)
    build_sheet_products(wb)
    build_sheet_plan_fact(wb)
    build_sheet_regions(wb)
    build_sheet_motivation(wb)
    build_sheet_cash_bank(wb)
    build_sheet_logistics(wb)
    build_sheet_dashboard(wb)

    wb.save(OUTPUT_PATH)
    print(f"Создан файл: {OUTPUT_PATH}")
    print("Листы: Сводка по годам, Продажи по месяцам, Топ товары, План-факт, Регионы, Мотивация ТП, Касса и банк, Логистика, Дашборд")


if __name__ == "__main__":
    main()
