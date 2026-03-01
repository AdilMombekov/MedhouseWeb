"""
Создание 7 шаблонов выгрузки для страниц дашборда.
Запуск: из папки backend: python -m scripts.create_dashboard_templates
"""
import sys
from pathlib import Path

backend = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(backend))

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

TEMPLATES_DIR = backend / "templates"
TEMPLATES_DIR.mkdir(exist_ok=True)

thin = Side(style="thin")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def header_row(ws, row):
    for c in range(1, ws.max_column + 1):
        ws.cell(row, c).font = Font(bold=True)
        ws.cell(row, c).border = border


def make_summary():
    wb = Workbook()
    ws = wb.active
    ws.title = "Сводка KPI"
    ws.append(["Период (год)", "Показатель", "Значение", "Кол-во записей", "Примечание"])
    header_row(ws, 1)
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 16
    wb.save(TEMPLATES_DIR / "dashboard_summary.xlsx")


def make_sales():
    wb = Workbook()
    ws = wb.active
    ws.title = "Продажи"
    ws.append(["Дата/Период", "Филиал/Город", "Сумма", "Количество", "Валюта"])
    header_row(ws, 1)
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 22
    wb.save(TEMPLATES_DIR / "dashboard_sales.xlsx")


def make_regions():
    wb = Workbook()
    ws = wb.active
    ws.title = "Регионы и филиалы"
    ws.append(["Код региона (KATO)", "Наименование", "Родитель (id)", "Показатель", "Значение"])
    header_row(ws, 1)
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 35
    wb.save(TEMPLATES_DIR / "dashboard_regions.xlsx")


def make_logistics():
    wb = Workbook()
    ws = wb.active
    ws.title = "Логистика"
    ws.append(["Дата", "Склад", "Товар/Номенклатура", "Приход", "Расход", "Остаток", "Ед. изм."])
    header_row(ws, 1)
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 35
    wb.save(TEMPLATES_DIR / "dashboard_logistics.xlsx")


def make_products():
    wb = Workbook()
    ws = wb.active
    ws.title = "Товарная номенклатура"
    ws.append(["Артикул", "Наименование", "Группа", "Количество", "Сумма", "Период"])
    header_row(ws, 1)
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 20
    wb.save(TEMPLATES_DIR / "dashboard_products.xlsx")


def make_plan_fact():
    wb = Workbook()
    ws = wb.active
    ws.title = "План-факт"
    ws.append(["Период", "План", "Факт", "Выполнение %", "Отклонение", "Комментарий"])
    header_row(ws, 1)
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["E"].width = 14
    wb.save(TEMPLATES_DIR / "dashboard_plan_fact.xlsx")


def make_executive():
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчёт для руководства"
    ws.append(["Показатель", "Значение", "Период", "Комментарий"])
    header_row(ws, 1)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["D"].width = 40
    wb.save(TEMPLATES_DIR / "dashboard_executive.xlsx")


def main():
    make_summary()
    make_sales()
    make_regions()
    make_logistics()
    make_products()
    make_plan_fact()
    make_executive()
    print("Созданы 7 файлов шаблонов в backend/templates/")

    from app.database import SessionLocal
    from app.models import ReportTemplate

    db = SessionLocal()
    types = [
        ("dashboard_summary", "dashboard_summary.xlsx", "Шаблон выгрузки: Сводка KPI", "Заполните по данным сводки и загрузите в раздел Импорт."),
        ("dashboard_sales", "dashboard_sales.xlsx", "Шаблон выгрузки: Продажи", "Динамика продаж по периодам и филиалам."),
        ("dashboard_regions", "dashboard_regions.xlsx", "Шаблон выгрузки: Регионы и филиалы", "Показатели по регионам KATO."),
        ("dashboard_logistics", "dashboard_logistics.xlsx", "Шаблон выгрузки: Логистика", "Приход, расход, остатки по складам."),
        ("dashboard_products", "dashboard_products.xlsx", "Шаблон выгрузки: Товарная номенклатура", "Продажи по товарам и группам."),
        ("dashboard_plan_fact", "dashboard_plan_fact.xlsx", "Шаблон выгрузки: План–факт", "Сравнение плана и факта по периодам."),
        ("dashboard_executive", "dashboard_executive.xlsx", "Шаблон выгрузки: Отчёт для руководства", "Сводные ключевые показатели за период."),
    ]
    for report_type, file_name, name, desc in types:
        if db.query(ReportTemplate).filter(ReportTemplate.report_type == report_type).first():
            continue
        db.add(ReportTemplate(name=name, description=desc, file_name=file_name, report_type=report_type))
    db.commit()
    db.close()
    print("Записи шаблонов добавлены в БД (если ещё не было).")


if __name__ == "__main__":
    main()
